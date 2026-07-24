from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .utils import json_value


def _value(cell: Any) -> Any:
    return json_value(cell.value)


def extract_workbook(path: Path) -> dict[str, Any]:
    formulas = load_workbook(path, data_only=False, read_only=False)
    values = load_workbook(path, data_only=True, read_only=False)
    sheets: dict[str, Any] = {}
    errors: list[dict[str, str]] = []
    for name in formulas.sheetnames:
        formula_sheet = formulas[name]
        value_sheet = values[name]
        rows: list[list[dict[str, Any]]] = []
        matrix: list[list[Any]] = []
        records: list[dict[str, Any]] = []
        headers: list[str] = []
        for row_number, formula_row in enumerate(formula_sheet.iter_rows(), 1):
            cells: list[dict[str, Any]] = []
            plain: list[Any] = []
            for formula_cell in formula_row:
                cached_cell = value_sheet[formula_cell.coordinate]
                formula_value = _value(formula_cell)
                cached_value = _value(cached_cell)
                if formula_value is None and cached_value is None:
                    plain.append(None)
                    continue
                item = {"cell": formula_cell.coordinate, "value": cached_value}
                if formula_cell.data_type == "f":
                    item["formula"] = formula_value
                cells.append(item)
                plain.append(cached_value if cached_value is not None else formula_value)
                for candidate in (formula_value, cached_value):
                    if isinstance(candidate, str) and candidate.startswith("#"):
                        errors.append({"sheet": name, "cell": formula_cell.coordinate, "value": candidate})
            if cells:
                rows.append(cells)
                matrix.append(plain)
                if not headers:
                    headers = [str(value or "") for value in plain]
                elif any(value not in (None, "") for value in plain):
                    records.append(
                        {
                            headers[index] or f"column_{index + 1}": value
                            for index, value in enumerate(plain)
                            if index < len(headers) and value not in (None, "")
                        }
                    )
        sheets[name] = {
            "dimensions": formula_sheet.calculate_dimension(),
            "hidden": formula_sheet.sheet_state != "visible",
            "rows": rows,
            "matrix": matrix,
            "records": records,
        }
    formulas.close()
    values.close()
    return {"source": str(path), "sheets": sheets, "formula_errors": errors}
